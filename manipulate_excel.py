import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import column_index_from_string
from typing import Optional, Union, Any


class ConnectionError(Exception):
    """Exceção customizada para estado inválido (arquivo ou planilha)."""
    pass


class Excel:
    """
    Classe para manipulação simplificada de arquivos Excel (.xlsx).

    Por padrão, seleciona a planilha ativa ao abrir.
    Use 'select_sheet(nome)' para escolher uma planilha diferente.
    """

    def __init__(self, path_file: str):
        """
        Inicializa o manipulador de Excel.

        :param path_file: Caminho para o arquivo .xlsx.
        """
        self.path_file = path_file
        self.workbook: Optional[Workbook] = None
        self.sheet: Optional[Worksheet] = None # Armazena a planilha selecionada

    def _check_workbook_is_open(self) -> None:
        """Verifica se o workbook (arquivo) foi carregado."""
        if not self.workbook:
            raise ConnectionError(
                "Arquivo Excel não está aberto. "
                "Chame o método 'open()' primeiro."
            )

    def _check_sheet_is_selected(self) -> None:
        """Verifica se uma planilha foi selecionada/carregada."""
        # Se 'self.sheet' for None, é porque 'open()' não foi chamado
        # ou 'close()' já foi chamado.
        if not self.sheet:
            raise ConnectionError(
                "Nenhuma planilha carregada. "
                "Chame o método 'open()' primeiro."
            )

    def open(self) -> None:
        """
        Abre o arquivo Excel (workbook). 
        
        Se o arquivo não existir, um novo será criado em memória.
        Automaticamente seleciona a planilha 'ativa' por padrão.
        """
        if self.workbook:
            return # Já está aberto
            
        try:
            self.workbook = openpyxl.load_workbook(self.path_file)
        except FileNotFoundError:
            self.workbook = Workbook()
        
        # --- COMPORTAMENTO PADRÃO ---
        # Seleciona a planilha ativa (geralmente a primeira) por padrão.
        self.sheet = self.workbook.active

    def close(self) -> None:
        """
        Salva o arquivo no disco e fecha o workbook, liberando recursos.
        """
        if self.workbook:
            try:
                self.workbook.save(self.path_file)
            except PermissionError:
                print(f"Erro: Permissão negada ao salvar '{self.path_file}'. "
                      "O arquivo pode estar aberto em outro programa.")
            finally:
                self.workbook.close()
                # Reseta o estado da classe
                self.workbook = None
                self.sheet = None

    def select_sheet(self, sheet_name: str) -> None:
        """
        Seleciona uma planilha (sheet) dentro do arquivo pelo nome.
        Todas as operações futuras usarão esta planilha.

        :param sheet_name: O nome da planilha a ser selecionada.
        :raises ValueError: Se a planilha com o nome especificado não for encontrada.
        """
        # Precisa que o workbook esteja aberto para selecionar uma planilha
        self._check_workbook_is_open()

        try:
            # Tenta acessar a planilha pelo nome
            self.sheet = self.workbook[sheet_name]

        except KeyError:
            # Se a planilha não existe, lança um erro
            # (self.workbook.sheetnames lista todas as planilhas existentes)
            raise ValueError(
                f"A planilha '{sheet_name}' não foi encontrada no arquivo. "
                f"Planilhas disponíveis: {self.workbook.sheetnames}"
            )

        except Exception as e:
            # Lida com outros erros inesperados
            raise ConnectionError(f"Erro ao selecionar planilha '{sheet_name}': {e}")

    def first_free_row(self, column: str) -> int:
        """
        Encontra a primeira linha vazia na planilha selecionada.

        :param column: A letra da coluna (ex: 'A', 'B', 'AA').
        :return: O número da primeira linha vazia (baseado em 1).
        """
        # Precisa que uma planilha esteja selecionada (ativa ou via select_sheet)
        self._check_sheet_is_selected()
        
        try:
            col_idx = column_index_from_string(column.upper())
        except ValueError:
            raise ValueError(f"Nome de coluna inválido: '{column}'")

        row = 1
        while self.sheet.cell(row=row, column=col_idx).value is not None:
            row += 1
            
        return row

    def add_value_to_cell(self, column: str, row: int, 
                          value: Union[str, float, int, None]) -> None:
        """
        Adiciona um valor a uma célula específica na planilha selecionada.

        :param column: A letra da coluna (ex: 'A').
        :param row: O número da linha (baseado em 1).
        :param value: O valor a ser inserido.
        """
        # Precisa que uma planilha esteja selecionada
        self._check_sheet_is_selected()
        
        try:
            col_idx = column_index_from_string(column.upper())
        except ValueError:
            raise ValueError(f"Nome de coluna inválido: '{column}'")
            
        if row < 1:
            raise ValueError("Número da linha (row) deve ser 1 ou maior.")

        self.sheet.cell(row=row, column=col_idx).value = value

    def get_cell_value(self, column: str, row: int) -> Any:
        """
        Lê e retorna o valor de uma célula específica na planilha selecionada.

        :param column: A letra da coluna (ex: 'A').
        :param row: O número da linha (baseado em 1).
        :return: O valor da célula (pode ser str, int, float, datetime, None, etc.).
        """
        # Precisa que uma planilha esteja selecionada
        self._check_sheet_is_selected()

        try:
            col_idx = column_index_from_string(column.upper())
        except ValueError:
            raise ValueError(f"Nome de coluna inválido: '{column}'")

        if row < 1:
            raise ValueError("Número da linha (row) deve ser 1 ou maior.")

        # Retorna o valor da célula
        return self.sheet.cell(row=row, column=col_idx).value

    # --- Métodos Mágicos (para 'with') ---
    def __enter__(self):
        """Permite o uso com a declaração 'with' (chama 'open()')."""
        self.open()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """Garante que o 'close()' (salvando) seja chamado ao sair do 'with'."""
        self.close()
